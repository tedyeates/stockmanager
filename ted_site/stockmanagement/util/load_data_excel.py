from collections import defaultdict
from decimal import Decimal
from openpyxl import load_workbook
from os.path import dirname
from stockmanagement.models import Group, Item, Instock, Brand, Outstock
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime

from ted_site.settings import BASE_DIR

import csv


DATA_LOC = dirname(dirname(__file__)) + "/static/stockmanagement/data/"


def load_stock_data():
    load_excel("stock_data.xlsx", is_instock=True,
                itemsheet="Tableitem", instocksheet="Instock", outstocksheet="OUTSTOCK21")


def load_excel(file_name, **kwargs):
    workbook = load_workbook(DATA_LOC + file_name)
    item_sheet = workbook[kwargs.pop("itemsheet")]
    instock_sheet = workbook[kwargs.pop("instocksheet")]
    outstock_sheet = workbook[kwargs.pop("outstocksheet")]
    open_log_file("invalid_item", Item, load_items, item_sheet)
    open_log_file("invalid_instock", Instock, load_instock, instock_sheet)
    open_log_file("invalid_outstock", Outstock, load_outstock, outstock_sheet)
    
    
def open_log_file(file_name, model,  function, *args, **kwargs):
    with open(f"{BASE_DIR}\stockmanagement\static\stockmanagement\{file_name}.csv", "w", newline="", encoding="utf-8") as csvfile:
        columns = [field.name for field in model._meta.get_fields()]

        writer = csv.DictWriter(csvfile, fieldnames=columns, restval="")
        writer.writeheader()
        function(writer, *args, **kwargs)
        

class Operation:
    DIVIDE = "/"
    MULTIPLY = "*"
    PLUS = "+"
    MINUS = "-"
    OPERATIONS = [DIVIDE, MULTIPLY, PLUS, MINUS]
    
def execute_calcultation(number1, operation, number2):
    if operation == Operation.DIVIDE:
        return Decimal(number1 / number2)
    if operation == Operation.MULTIPLY:
        return Decimal(number1 * number2)
    if operation == Operation.PLUS:
        return Decimal(number1 + number2)
    if operation == Operation.MINUS:
        return Decimal(number1 - number2)
        
        
def calculate_excel_function(function_string):
    print(function_string)
    # Remove equal sign
    equation = function_string[1:]
    
    equation_split = defaultdict(str)
    position = 0
    for character in equation:
        if character not in Operation.OPERATIONS:
            equation_split[position] += character
            continue
        
        equation_split[position + 1] = character
        position += 2
        
    print(equation_split)
    
    for operation in Operation.OPERATIONS:
        if operation not in equation: continue
        
        index = 0
        while index < len(equation_split):
            if equation_split[index] == operation:
                first_number = Decimal(equation_split[index-1])
                second_number = Decimal(equation_split[index+1])
                equation_split[index-1] = execute_calcultation(first_number, operation, second_number)
                del equation_split[index]
                del equation_split[index+1]
                index -= 1
            index += 1
            
    return equation_split[0]
    

def get_cell_value(row, column, is_date=False, is_upper=False):
    cell = row[column].value
    if cell is None: return None
    
    if isinstance(cell, datetime): return cell
    if is_date: return datetime.strptime(cell, '%d/%m/%y')
    if is_upper: return cell.upper()
    
    if isinstance(cell, str) and cell.isdecimal():
        cell = Decimal(cell)
    
    if not isinstance(cell, str) or cell.isdecimal(): 
        return Decimal(round(cell,2))
    
    if isinstance(cell, str) and cell[0] == "=":
        return calculate_excel_function(cell)
    
    return cell.strip()
    
    
def load_items(writer, item_sheet):
    ITEM_CODE_COL = 1
    ITEM_GROUP_COL = 2
    ITEM_DESC_COL = 3
    ITEM_BRAND_COL = 4
    ITEM_UNIT_COL = 5
    ITEM_WEIGHT_COL = 10

    for row in item_sheet.iter_rows(min_row=3):
        item_code = get_cell_value(row, ITEM_CODE_COL, is_upper=True)

        item_desc = get_cell_value(row, ITEM_DESC_COL)
        item_brand = get_cell_value(row, ITEM_BRAND_COL)
        item_unit = get_cell_value(row, ITEM_UNIT_COL)
        item_group = get_cell_value(row, ITEM_GROUP_COL)
        item_weight = get_cell_value(row, ITEM_WEIGHT_COL)

        formatted_row = {
            "code": item_code, "brand": item_brand, "unit": item_unit, 
            "group": item_group, "weight": item_weight
        }
        
        if item_code == None:
            writer.writerow(formatted_row)
            continue
        
        group = None
        if item_group:
            item_group = item_group.strip()
            if item_group != "":
                group, _ = Group.objects.get_or_create(name=item_group)
            
            
        brand = None
        if item_brand:
            item_brand = item_brand.strip()
            brand, _ = Brand.objects.get_or_create(name=item_brand)
                
        if isinstance(item_weight, str) and not item_weight.isdecimal():
            item_weight = None
        
        Item.objects.get_or_create(code=item_code, defaults={"description": item_desc,
                                                            "brand": brand,
                                                            "unit": item_unit,
                                                            "group": group,
                                                            "weight": item_weight})

def get_min_price(stock_price, item):
    if item.min_price == -1 or item.min_price is None:
        return stock_price
    if stock_price is None:
        return item.min_price
    return min(item.min_price, stock_price)

def get_max_price(stock_price, item):
    if item.max_price == -1 or item.max_price is None:
        return stock_price
    if stock_price is None:
        return item.max_price
    return max(item.max_price, stock_price)


def load_instock(writer, stock_sheet):
    DATE_COL = 0
    IV_COL = 1
    PO_COL = 2
    PC_COL = 3
    SUPPLIER_COL = 4
    ITEM_COL = 5
    QUANTITY_COL = 6
    PRICE_COL = 7

    count = 0
    for row in stock_sheet.iter_rows(min_row=3):
        count+=1
        stock_iv = get_cell_value(row, IV_COL)
        stock_item = get_cell_value(row, ITEM_COL, is_upper=True)
        stock_date = get_cell_value(row, DATE_COL, is_date=True)
        stock_po =  get_cell_value(row, PO_COL)
        stock_pc = get_cell_value(row, PC_COL)
        stock_supplier = get_cell_value(row, SUPPLIER_COL)
        stock_quantity = get_cell_value(row, QUANTITY_COL)
        stock_price = get_cell_value(row, PRICE_COL)
            
        formatted_row = {
            "invoice_id": stock_iv, "item": stock_item, "quantity": stock_quantity,
            "stock_date": stock_date, "purchase_order_id": stock_po, "job_id": stock_pc,
            "supplier": stock_supplier, "price": stock_price
        }
        should_skip = False
        for cell in [stock_iv, stock_item]:
            if cell is None or cell == "":
                should_skip = True
                
        if should_skip: 
            writer.writerow(formatted_row)
            continue
        
            
        if stock_quantity is None or stock_quantity == "":
            print("[CRIT] No Quantity")
            writer.writerow(formatted_row)
            continue
        
        if stock_price is None or stock_price == "":
            print("[CRIT] No Price")
            writer.writerow(formatted_row)
            continue
        
        if (isinstance(stock_quantity, str) and not stock_quantity.isdecimal()) or (isinstance(stock_price, str) and not stock_price.isdecimal()):
            writer.writerow(formatted_row)
            continue

        try:
            item = Item.objects.get(code=stock_item)
        except ObjectDoesNotExist:
            print(f"[CRIT] No Matching Item {stock_item}, Stock {stock_iv} not added to database {stock_price} {stock_pc} {stock_po}")
            writer.writerow(formatted_row)
            continue

        if isinstance(stock_price, str):
            stock_price.replace(" ", "")

        if isinstance(stock_quantity, str):
            stock_quantity.replace(" ", "")
        
        total_price = Decimal(stock_price) * Decimal(stock_quantity)
        item.min_price = get_min_price(stock_price, item)
        item.max_price = get_max_price(stock_price, item)
        item.sum_price = sum([Decimal(item.sum_price), total_price])
        item.quantity += stock_quantity
        item.instock_number += Decimal(1)
        item.save()
            
        Instock.objects.get_or_create(
            invoice_id=stock_iv, item=item, purchase_order_id=stock_po, 
            job_id=stock_pc,
            defaults={
                "stock_date": stock_date,
                "supplier": stock_supplier,
                "quantity": stock_quantity,
                "price": stock_price
            }
        )


def load_outstock(writer, outstock_sheet):
    DATE_COL = 0
    PC_COL = 1
    CUSTOMER_COL = 2
    STOCK_COL = 3
    REQUESTER_COL = 4
    DEPARTMENT_COL = 5
    ITEM_COL = 6
    QUANTITY_COL = 7
    
    count = 0
    for row in outstock_sheet.iter_rows(min_row=3):
        count+=1
        
        stock_pc = get_cell_value(row, PC_COL)
        stock =  get_cell_value(row, STOCK_COL)
        stock_item = get_cell_value(row, ITEM_COL, is_upper=True)
        stock_date = get_cell_value(row, DATE_COL, is_date=True)
        stock_requester = get_cell_value(row, REQUESTER_COL)
        stock_quantity = get_cell_value(row, QUANTITY_COL)
        stock_department = get_cell_value(row, DEPARTMENT_COL)
        stock_customer = get_cell_value(row, CUSTOMER_COL)
        
        formatted_row = {
            "job_id": stock_pc, "stock_date": stock_date, "requester": stock_requester,
            "stock_id": stock, "item": stock_item, "quantity": stock_quantity, 
            "department": stock_department, "customer": stock_customer
        }
        
        should_skip = False
        for cell in [stock_pc, stock, stock_item]:
            if cell is None or cell == "":
                should_skip = True
                
        if should_skip: 
            writer.writerow(formatted_row)
            continue
        
            
        if stock_quantity is None or stock_quantity == "":
            print("[CRIT] No Quantity")
            writer.writerow(formatted_row)
            continue
        
        if (isinstance(stock_quantity, str) and not stock_quantity.isdecimal()):
            writer.writerow(formatted_row)
            continue

        try:
            item = Item.objects.get(code=stock_item)

            if isinstance(stock_quantity, str):
                stock_quantity.replace(" ", "")
                
            item.quantity -= Decimal(stock_quantity)
            if item.quantity < 0:
                item.quantity = 0
            
            item.outstock_number += Decimal(1)
            item.save()
                
            Outstock.objects.get_or_create(stock_id=stock, job_id=stock_pc, item=item,
                                        defaults={"stock_date": stock_date,
                                                  "requester": stock_requester,
                                                  "quantity": stock_quantity,
                                                  "department": stock_department,
                                                  "customer": stock_customer })
        except ObjectDoesNotExist:
            print(f"[CRIT] No Matching Item {item.code}, Stock {stock} not added to database")
            writer.writerow(formatted_row)
    