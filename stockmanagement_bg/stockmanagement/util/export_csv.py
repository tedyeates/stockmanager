"""
Generate CSVs from Excel stock data for fast import via psql \copy.

Assumes reset_django_tables has been run — IDs are assigned sequentially starting at 1.

Usage:
    python manage.py loadstock --export-csv

Then import with:
    psql <conn> -c "\copy stockmanagement_group FROM 'groups.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_brand FROM 'brands.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_item FROM 'items.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_instock FROM 'instock.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_outstock FROM 'outstock.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_instock_job FROM 'instock_jobs.csv' WITH CSV HEADER"
    psql <conn> -c "\copy stockmanagement_outstock_job FROM 'outstock_jobs.csv' WITH CSV HEADER"
    psql <conn> -c "\copy job FROM 'jobs.csv' WITH CSV HEADER"
    psql <conn> -c "\copy customer FROM 'customers.csv' WITH CSV HEADER"
"""
import csv
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from stockmanagement.models import StoreType

DATA_LOC = Path(__file__).resolve().parent.parent / "static" / "stockmanagement" / "data"
OUTPUT_DIR = DATA_LOC / "csv_export"


# ---------------------------------------------------------------------------
# Cell helpers (same as load_data_excel)
# ---------------------------------------------------------------------------

def get_value(row, col, is_upper=False):
    val = row[col]
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.startswith("#"):
            return None
        return val.upper() if is_upper else val
    return val


def safe_decimal(value):
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def clean_job_id(job_val):
    if job_val is None:
        return None
    job_str = str(job_val).strip()
    if not job_str or job_str == '.':
        return None
    if '.' in job_str:
        try:
            job_str = str(int(float(job_str)))
        except (ValueError, TypeError):
            pass
    return job_str or None


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def export_stock_csvs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Shared state across both files
    state = ExportState()

    process_excel(
        state, "ACC_2026.xlsx",
        itemsheet="TABLE ITEM", instocksheet="Instock", outstocksheet="Outstock",
        store_type=StoreType.ACCESSORY, has_group_col=False, outstock_header_row=1,
    )
    process_excel(
        state, "METAL_2026.xlsx",
        itemsheet="Tableitem", instocksheet="Instock", outstocksheet="Outstock.",
        store_type=StoreType.METAL, has_group_col=True, outstock_header_row=2,
    )

    state.write_all()
    print(f"[DONE] CSVs written to {OUTPUT_DIR}")


# ---------------------------------------------------------------------------
# State tracker — assigns IDs and accumulates rows
# ---------------------------------------------------------------------------

class ExportState:
    def __init__(self):
        self.groups = {}       # name -> id
        self.brands = {}       # name -> id
        self.items = {}        # code -> id
        self.jobs = {}         # job_id_str -> job_id_str (just tracking existence)
        self.customers = {}    # name -> id

        self.group_rows = []
        self.brand_rows = []
        self.item_rows = []
        self.instock_rows = []
        self.outstock_rows = []
        self.instock_job_rows = []
        self.outstock_job_rows = []
        self.job_rows = []
        self.customer_rows = []

        self._next_group_id = 1
        self._next_brand_id = 1
        self._next_item_id = 1
        self._next_instock_id = 1
        self._next_outstock_id = 1
        self._next_instock_job_id = 1
        self._next_outstock_job_id = 1
        self._next_customer_id = 1

        # Item accumulators for price/quantity
        self._item_data = {}  # code -> dict of computed fields

    def get_or_create_group(self, name):
        if name in self.groups:
            return self.groups[name]
        gid = self._next_group_id
        self._next_group_id += 1
        self.groups[name] = gid
        now = datetime.now().isoformat()
        self.group_rows.append([gid, now, name, ""])
        return gid

    def get_or_create_brand(self, name):
        if name in self.brands:
            return self.brands[name]
        bid = self._next_brand_id
        self._next_brand_id += 1
        self.brands[name] = bid
        now = datetime.now().isoformat()
        self.brand_rows.append([bid, now, name])
        return bid

    def get_or_create_job(self, job_id_str, customer_id=None):
        if job_id_str in self.jobs:
            # Update customer if we now have one and didn't before
            if customer_id and not self.jobs[job_id_str]:
                self.jobs[job_id_str] = customer_id
                # Update the existing row
                for row in self.job_rows:
                    if row[0] == job_id_str:
                        row[1] = customer_id
                        break
            return job_id_str
        self.jobs[job_id_str] = customer_id
        self.job_rows.append([job_id_str, customer_id])
        return job_id_str

    def get_or_create_customer(self, name):
        if name in self.customers:
            return self.customers[name]
        cid = self._next_customer_id
        self._next_customer_id += 1
        self.customers[name] = cid
        self.customer_rows.append([cid, name])
        return cid

    def create_item(self, code, description, brand_id, unit, group_id, min_qty, max_qty):
        if code in self.items:
            return self.items[code]
        item_id = self._next_item_id
        self._next_item_id += 1
        self.items[code] = item_id
        self._item_data[code] = {
            "quantity": Decimal(0),
            "instock_number": 0,
            "outstock_number": 0,
            "sum_price": Decimal(0),
            "min_price": None,
            "max_price": None,
        }
        now = datetime.now().isoformat()
        self.item_rows.append({
            "id": item_id, "modified": now, "code": code,
            "description": description or "", "unit": unit or None,
            "instock_number": 0, "outstock_number": 0,
            "max_price": None, "sum_price": 0, "min_price": None,
            "quantity": 0, "notes": None,
            "min_quanity": min_qty or None, "max_quanity": max_qty or None,
            "brand_id": brand_id or None, "group_id": group_id or None,
        })
        return item_id

    def add_instock(self, item_code, stock_date, invoice_id, po, supplier, quantity, price, store_type, job_str):
        item_id = self.items.get(item_code)
        if item_id is None:
            print(f"[CRIT] No item {item_code} for instock")
            return

        instock_id = self._next_instock_id
        self._next_instock_id += 1
        now = datetime.now().isoformat()
        date_str = stock_date.strftime("%Y-%m-%d") if isinstance(stock_date, datetime) else (stock_date or None)

        self.instock_rows.append([
            instock_id, date_str, now, now,
            str(quantity or 0), None, store_type,
            str(invoice_id), str(price) if price is not None else None, po or None, supplier or None,
            item_id,
        ])

        # Update item aggregates
        data = self._item_data[item_code]
        if quantity is not None and price is not None:
            data["instock_number"] += 1
            data["quantity"] += quantity
            data["sum_price"] += price * quantity
            if data["min_price"] is None or price < data["min_price"]:
                data["min_price"] = price
            if data["max_price"] is None or price > data["max_price"]:
                data["max_price"] = price
        elif quantity is not None:
            data["quantity"] += quantity

        # M2M job
        if job_str:
            job_id = clean_job_id(job_str)
            if job_id:
                self.get_or_create_job(job_id)
                m2m_id = self._next_instock_job_id
                self._next_instock_job_id += 1
                self.instock_job_rows.append([m2m_id, instock_id, job_id])

    def add_outstock(self, item_code, stock_date, stock_id, requester, dept, quantity, store_type, job_str, customer_name):
        item_id = self.items.get(item_code)
        if item_id is None:
            print(f"[CRIT] No item {item_code} for outstock")
            return

        outstock_id = self._next_outstock_id
        self._next_outstock_id += 1
        now = datetime.now().isoformat()
        date_str = stock_date.strftime("%Y-%m-%d") if isinstance(stock_date, datetime) else (stock_date or None)

        customer_id = None
        if customer_name:
            customer_id = self.get_or_create_customer(str(customer_name))

        # Update item aggregates
        data = self._item_data[item_code]
        if quantity is not None:
            data["quantity"] = max(Decimal(0), data["quantity"] - quantity)
            data["outstock_number"] += 1

        remaining = data["quantity"]

        self.outstock_rows.append([
            outstock_id, date_str, now, now,
            str(quantity or 0), None, store_type,
            stock_id or None, requester or "", dept or None,
            str(remaining), item_id, customer_id,
        ])

        # M2M job
        if job_str:
            job_id = clean_job_id(job_str)
            if job_id:
                self.get_or_create_job(job_id, customer_id or None)
                m2m_id = self._next_outstock_job_id
                self._next_outstock_job_id += 1
                self.outstock_job_rows.append([m2m_id, outstock_id, job_id])

    def write_all(self):
        # Finalize item rows with computed aggregates
        for row in self.item_rows:
            code = row["code"]
            data = self._item_data[code]
            row["quantity"] = str(data["quantity"])
            row["instock_number"] = data["instock_number"]
            row["outstock_number"] = data["outstock_number"]
            row["sum_price"] = str(data["sum_price"])
            row["min_price"] = str(data["min_price"]) if data["min_price"] is not None else None
            row["max_price"] = str(data["max_price"]) if data["max_price"] is not None else None

        # Groups
        _write_csv("groups.csv",
                   ["id", "modified", "name", "description"],
                   self.group_rows)

        # Brands
        _write_csv("brands.csv",
                   ["id", "modified", "name"],
                   self.brand_rows)

        # Items
        item_cols = ["id", "modified", "code", "description", "unit",
                     "instock_number", "outstock_number", "max_price", "sum_price",
                     "min_price", "quantity", "notes", "min_quanity", "max_quanity",
                     "brand_id", "group_id"]
        with open(OUTPUT_DIR / "items.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=item_cols)
            writer.writeheader()
            for row in self.item_rows:
                writer.writerow({k: _null(v) for k, v in row.items()})
        print(f"  items.csv: {len(self.item_rows)} rows")

        # Instock
        _write_csv("instock.csv",
                   ["id", "stock_date", "created_date", "modified",
                    "quantity", "notes", "store_type", "invoice_id", "price",
                    "purchase_order_id", "supplier", "item_id"],
                   self.instock_rows)

        # Outstock
        _write_csv("outstock.csv",
                   ["id", "stock_date", "created_date", "modified",
                    "quantity", "notes", "store_type", "stock_id", "requester",
                    "department", "remaining_quantity", "item_id", "customer_id"],
                   self.outstock_rows)

        # M2M tables
        _write_csv("instock_jobs.csv",
                   ["id", "instock_id", "job_id"],
                   self.instock_job_rows)

        _write_csv("outstock_jobs.csv",
                   ["id", "outstock_id", "job_id"],
                   self.outstock_job_rows)

        # Unmanaged tables
        _write_csv("jobs.csv", ["job_id", "customer_id"], self.job_rows)
        _write_csv("customers.csv", ["id", "name"], self.customer_rows)


def _write_csv(filename, headers, rows):
    path = OUTPUT_DIR / filename
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([_null(v) for v in row])
    print(f"  {filename}: {len(rows)} rows")


def _null(value):
    """Convert None to \\N for psql COPY NULL handling."""
    if value is None:
        return r"\N"
    return value


# ---------------------------------------------------------------------------
# Excel processing
# ---------------------------------------------------------------------------

def process_excel(state, file_name, **kwargs):
    file_path = DATA_LOC / file_name
    print(f"Processing {file_path}")
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    store_type = kwargs["store_type"]
    has_group_col = kwargs.get("has_group_col", False)
    outstock_header_row = kwargs.get("outstock_header_row", 1)

    process_items(state, workbook[kwargs["itemsheet"]], store_type, has_group_col)
    process_instock(state, workbook[kwargs["instocksheet"]], store_type)
    process_outstock(state, workbook[kwargs["outstocksheet"]], store_type, outstock_header_row)

    workbook.close()


def process_items(state, sheet, store_type, has_group_col):
    COL = (
        dict(code=0, group=1, desc=2, brand=3, unit=4, max_price=5, min_price=6)
        if has_group_col else
        dict(code=0, max_qty=1, min_qty=2, desc=3, brand=4, unit=5, max_price=6, min_price=7)
    )

    for row_vals in sheet.iter_rows(min_row=2, values_only=True):
        code = get_value(row_vals, COL["code"], is_upper=True)
        if not code:
            continue
        code = str(code)
        if code in state.items:
            continue

        desc = get_value(row_vals, COL["desc"])
        brand_val = get_value(row_vals, COL["brand"])
        unit = get_value(row_vals, COL["unit"])
        group_val = get_value(row_vals, COL["group"]) if has_group_col else None
        max_qty = safe_decimal(get_value(row_vals, COL["max_qty"])) if "max_qty" in COL else None
        min_qty = safe_decimal(get_value(row_vals, COL["min_qty"])) if "min_qty" in COL else None

        group_id = None
        if group_val:
            group_id = state.get_or_create_group(str(group_val).strip())

        brand_id = None
        if brand_val:
            brand_id = state.get_or_create_brand(str(brand_val).strip())

        state.create_item(code, desc, brand_id, unit, group_id, min_qty, max_qty)

    print(f"  Items processed: {len(state.items)} total")


def process_instock(state, sheet, store_type):
    DATE_COL, IV_COL, PO_COL, PC_COL = 0, 1, 2, 3
    SUPPLIER_COL, ITEM_COL, QTY_COL, PRICE_COL = 4, 5, 6, 7
    count = 0

    for row_vals in sheet.iter_rows(min_row=2, values_only=True):
        iv = get_value(row_vals, IV_COL)
        item_id = get_value(row_vals, ITEM_COL, is_upper=True)
        if not iv or not item_id:
            continue

        item_id = str(item_id)
        date = get_value(row_vals, DATE_COL)
        po = get_value(row_vals, PO_COL)
        pc = get_value(row_vals, PC_COL)
        supplier = get_value(row_vals, SUPPLIER_COL)
        quantity = safe_decimal(get_value(row_vals, QTY_COL))
        price = safe_decimal(get_value(row_vals, PRICE_COL))

        job_str = str(pc) if pc is not None else None

        state.add_instock(item_id, date, iv, po, supplier, quantity, price, store_type, job_str)
        count += 1

    print(f"  Instock rows: {count}")


def process_outstock(state, sheet, store_type, header_row):
    DATE_COL, JOB_COL, CUST_COL, ST_COL = 0, 1, 2, 3
    REQ_COL, DEPT_COL, ITEM_COL, QTY_COL = 4, 5, 6, 7
    count = 0

    for row_vals in sheet.iter_rows(min_row=header_row + 2, values_only=True):
        job_val = get_value(row_vals, JOB_COL)
        stock_id = get_value(row_vals, ST_COL)
        item_id = get_value(row_vals, ITEM_COL, is_upper=True)
        if not job_val or not item_id:
            continue

        item_id = str(item_id)
        date = get_value(row_vals, DATE_COL)
        cust_val = get_value(row_vals, CUST_COL)
        requester = get_value(row_vals, REQ_COL)
        dept = get_value(row_vals, DEPT_COL)
        quantity = safe_decimal(get_value(row_vals, QTY_COL))

        state.add_outstock(item_id, date, stock_id, requester, dept, quantity, store_type, str(job_val), cust_val)
        count += 1

    print(f"  Outstock rows: {count}")
