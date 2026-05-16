from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from stockmanagement.models import Customer, Group, Item, Instock, Brand, Outstock, StoreType, Job
from datetime import datetime
from pathlib import Path
from stockmanagement_bg.settings import BASE_DIR
import csv
import json

DATA_LOC = Path(__file__).resolve().parent.parent / "static" / "stockmanagement" / "data"
PROGRESS_FILE = DATA_LOC / "progress.json"

# How many rows to accumulate before flushing to DB
FLUSH_EVERY = 200


# ---------------------------------------------------------------------------
# Progress tracking
# ---------------------------------------------------------------------------

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, "r") as f:
            progress = json.load(f)
            print(f"[RESUME] Resuming from: {progress}")
            return progress
    return {}


def save_progress(progress):
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def clear_progress():
    if PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink()
        print("[INFO] All sheets complete — progress file cleared")


def progress_key(file_name, sheet_name):
    return f"{file_name}::{sheet_name}"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def load_stock_data():
    load_excel(
        "ACC_2026.xlsx",
        itemsheet="TABLE ITEM",
        instocksheet="Instock",
        outstocksheet="Outstock",
        store_type=StoreType.ACCESSORY,
        has_group_col=False,
        outstock_header_row=1,
    )
    load_excel(
        "METAL_2026.xlsx",
        itemsheet="Tableitem",
        instocksheet="Instock",
        outstocksheet="Outstock.",
        store_type=StoreType.METAL,
        has_group_col=True,
        outstock_header_row=2,
    )
    clear_progress()


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def load_excel(file_name, **kwargs):
    file_path = DATA_LOC / file_name
    print(f"Loading {file_path} (exists: {file_path.exists()})")
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    store_type          = kwargs["store_type"]
    has_group_col       = kwargs.get("has_group_col", False)
    outstock_header_row = kwargs.get("outstock_header_row", 1)

    open_log_file("invalid_item",     Item,     load_items,    workbook[kwargs["itemsheet"]],     file_name, kwargs["itemsheet"],     store_type, has_group_col)
    open_log_file("invalid_instock",  Instock,  load_instock,  workbook[kwargs["instocksheet"]],  file_name, kwargs["instocksheet"],  store_type)
    open_log_file("invalid_outstock", Outstock, load_outstock, workbook[kwargs["outstocksheet"]], file_name, kwargs["outstocksheet"], store_type, outstock_header_row)

    workbook.close()


def open_log_file(log_name, model, function, sheet, *args):
    log_path = BASE_DIR / "stockmanagement" / "static" / "stockmanagement" / f"{log_name}.csv"
    columns = [field.name for field in model._meta.get_fields()] + ["source_file", "row_number", "notes"]
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with open(log_path, "a", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=columns, restval="")
        if write_header:
            writer.writeheader()
        function(writer, sheet, *args)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def get_value(row, col, is_upper=False):
    val = row[col]
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.startswith("#"):
            return None
        return val.upper() if is_upper else val
    return val


def safe_decimal(value):
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Lookup caches
# ---------------------------------------------------------------------------

def build_caches():
    return {
        "items":     {i.code: i for i in Item.objects.all()},
        "groups":    {g.name: g for g in Group.objects.all()},
        "brands":    {b.name: b for b in Brand.objects.all()},
        "customers": {c.name: c for c in Customer.objects.all()},
    }


def get_or_create_cached(cache, model, key_field, key_value, **defaults):
    if key_value in cache:
        return cache[key_value], False
    obj, created = model.objects.get_or_create(**{key_field: key_value}, defaults=defaults)
    cache[key_value] = obj
    return obj, created


# ---------------------------------------------------------------------------
# Flush helpers
# ---------------------------------------------------------------------------

def flush_items(items_by_code, fields):
    """Bulk update deduped item dict and clear it."""
    if not items_by_code:
        return
    Item.objects.bulk_update(list(items_by_code.values()), fields, batch_size=500)
    print(f"[INFO] Flushed {len(items_by_code)} item updates")
    items_by_code.clear()


def flush_stock_records(pending, model):
    """Bulk create stock records and return created objects."""
    if not pending:
        return []
    created = model.objects.bulk_create(pending, batch_size=500)
    print(f"[INFO] Bulk created {len(created)} {model.__name__} records")
    pending.clear()
    return created



# ---------------------------------------------------------------------------
# Items
# ---------------------------------------------------------------------------

def load_items(writer, sheet, file_name, sheet_name, store_type, has_group_col):
    progress    = load_progress()
    key         = progress_key(file_name, sheet_name)
    resume_from = progress.get(key, 0)

    if resume_from:
        print(f"[RESUME] Skipping to row {resume_from} for {key}")

    COL = (
        dict(code=0, group=1, desc=2, brand=3, unit=4, max_price=5, min_price=6)
        if has_group_col else
        dict(code=0, max_qty=1, min_qty=2, desc=3, brand=4, unit=5, max_price=6, min_price=7)
    )

    caches        = build_caches()
    pending_items = []  # unsaved Item objects
    pending_codes = {}  # code -> index in pending_items, for dedup

    def flush_pending_items():
        """Bulk create pending items then reload them into cache with DB-assigned PKs."""
        if not pending_items:
            return
        Item.objects.bulk_create(pending_items, ignore_conflicts=True, batch_size=500)
        # Re-query to get PKs — needed so instock/outstock FKs resolve correctly
        for item in Item.objects.filter(code__in=list(pending_codes.keys())):
            caches["items"][item.code] = item
        print(f"[INFO] Bulk created {len(pending_items)} items")
        pending_items.clear()
        pending_codes.clear()

    for row_num, row_vals in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row_num < resume_from:
            continue

        code = get_value(row_vals, COL["code"], is_upper=True)
        if not code:
            print(f"[SKIP] Items row {row_num + 2} — empty code")
            writer.writerow({"code": "", "source_file": file_name, "row_number": row_num + 2, "notes": "empty code"})
            continue
        code = str(code)

        desc      = get_value(row_vals, COL["desc"])
        brand_val = get_value(row_vals, COL["brand"])
        unit      = get_value(row_vals, COL["unit"])
        group_val = get_value(row_vals, COL["group"]) if has_group_col else None
        max_qty   = safe_decimal(get_value(row_vals, COL["max_qty"])) if "max_qty" in COL else None
        min_qty   = safe_decimal(get_value(row_vals, COL["min_qty"])) if "min_qty" in COL else None

        group = None
        if group_val:
            group, _ = get_or_create_cached(caches["groups"], Group, "name", str(group_val).strip())

        brand = None
        if brand_val:
            brand, _ = get_or_create_cached(caches["brands"], Brand, "name", str(brand_val).strip())

        # Skip codes already in DB or already queued
        if code in caches["items"] or code in pending_codes:
            print(f"[SKIP] Items row {row_num + 2} — duplicate code {code}")
            writer.writerow({"code": code, "source_file": file_name, "row_number": row_num + 2, "notes": "duplicate code"})
        else:
            pending_codes[code] = len(pending_items)
            pending_items.append(Item(
                code=code, description=desc or "",
                brand=brand, unit=unit, group=group,
                min_quanity=min_qty, max_quanity=max_qty,
            ))

        if len(pending_items) >= FLUSH_EVERY:
            flush_pending_items()
            progress[key] = row_num + 1
            save_progress(progress)
            print(f"[PROGRESS] {sheet_name} — row {row_num + 1}")

    flush_pending_items()
    progress.pop(key, None)
    save_progress(progress)
    print(f"[INFO] Items sheet complete: {sheet_name}")


# ---------------------------------------------------------------------------
# Instock
# ---------------------------------------------------------------------------

def load_instock(writer, sheet, file_name, sheet_name, store_type):
    progress    = load_progress()
    key         = progress_key(file_name, sheet_name)
    resume_from = progress.get(key, 0)

    if resume_from:
        print(f"[RESUME] Skipping to row {resume_from} for {key}")

    DATE_COL, IV_COL, PO_COL, PC_COL            = 0, 1, 2, 3
    SUPPLIER_COL, ITEM_COL, QTY_COL, PRICE_COL  = 4, 5, 6, 7
    ITEM_FIELDS = ["min_price", "max_price", "sum_price", "quantity", "instock_number"]

    caches          = build_caches()
    items_by_code   = {}   # deduped: only the latest state of each item
    pending_instock = []   # batch of Instock objects to bulk_create
    pending_jobs    = []   # parallel list of job_id strings for M2M assignment

    for row_num, row_vals in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row_num < resume_from:
            continue

        iv      = get_value(row_vals, IV_COL)
        item_id = get_value(row_vals, ITEM_COL, is_upper=True)

        if not iv or not item_id:
            print(f"[SKIP] Instock row {row_num + 2} — missing iv={iv} or item_id={item_id}")
            writer.writerow({"invoice_id": iv or "", "item": item_id or "", "source_file": file_name, "row_number": row_num + 2, "notes": "missing iv or item_id"})
            continue

        item_id  = str(item_id)
        date     = get_value(row_vals, DATE_COL)
        po       = get_value(row_vals, PO_COL)
        pc       = get_value(row_vals, PC_COL)
        supplier = get_value(row_vals, SUPPLIER_COL)
        quantity = safe_decimal(get_value(row_vals, QTY_COL))
        price    = safe_decimal(get_value(row_vals, PRICE_COL))

        if quantity is None or price is None:
            print(f"[WARN] Missing qty/price — IV {iv} item {item_id} (row saved with nulls)")

        item = caches["items"].get(item_id)
        if item is None:
            print(f"[CRIT] No matching item {item_id} for instock {iv}")
            writer.writerow({"invoice_id": iv, "item": item_id, "source_file": file_name, "row_number": row_num + 2, "notes": "no matching item in DB"})
            continue

        job_str = str(pc) if pc is not None else None

        # Accumulate item aggregate changes in memory — only when qty and price present
        tracked = items_by_code.get(item_id, item)
        if price is not None:
            tracked.min_price = _min_price(price, tracked)
            tracked.max_price = _max_price(price, tracked)
        if quantity is not None:
            tracked.quantity = (tracked.quantity or Decimal(0)) + quantity
        if quantity is not None and price is not None:
            total = Decimal(price) * Decimal(quantity)
            tracked.sum_price = (tracked.sum_price or Decimal(0)) + total
            tracked.instock_number += 1
        items_by_code[item_id] = tracked

        pending_instock.append(Instock(
            invoice_id=str(iv), item=item, purchase_order_id=po,
            stock_date=date, supplier=supplier,
            quantity=quantity or Decimal(0), price=price,
            store_type=store_type,
        ))
        pending_jobs.append(job_str)

        if len(pending_instock) >= FLUSH_EVERY:
            created = flush_stock_records(pending_instock, Instock)
            _assign_jobs(created, pending_jobs)
            pending_jobs.clear()
            flush_items(items_by_code, ITEM_FIELDS)
            progress[key] = row_num + 1
            save_progress(progress)
            print(f"[PROGRESS] {sheet_name} — row {row_num + 1}")

    # Final flush
    created = flush_stock_records(pending_instock, Instock)
    _assign_jobs(created, pending_jobs)
    pending_jobs.clear()
    flush_items(items_by_code, ITEM_FIELDS)
    progress.pop(key, None)
    save_progress(progress)
    print(f"[INFO] Instock sheet complete: {sheet_name}")


# ---------------------------------------------------------------------------
# Outstock
# ---------------------------------------------------------------------------

def load_outstock(writer, sheet, file_name, sheet_name, store_type, header_row):
    progress    = load_progress()
    key         = progress_key(file_name, sheet_name)
    resume_from = progress.get(key, 0)

    if resume_from:
        print(f"[RESUME] Skipping to row {resume_from} for {key}")

    DATE_COL, JOB_COL, CUST_COL, ST_COL  = 0, 1, 2, 3
    REQ_COL, DEPT_COL, ITEM_COL, QTY_COL = 4, 5, 6, 7
    ITEM_FIELDS = ["quantity", "outstock_number"]

    caches           = build_caches()
    items_by_code    = {}
    pending_outstock = []
    pending_jobs     = []   # parallel list of job_id strings for M2M assignment

    for row_num, row_vals in enumerate(sheet.iter_rows(min_row=header_row + 2, values_only=True)):
        if row_num < resume_from:
            continue

        job_val  = get_value(row_vals, JOB_COL)
        stock_id = get_value(row_vals, ST_COL)
        item_id  = get_value(row_vals, ITEM_COL, is_upper=True)

        if not job_val or not item_id:
            print(f"[SKIP] Outstock row {row_num + 2} — missing job={job_val} item_id={item_id}")
            writer.writerow({"stock_id": stock_id or "", "item": item_id or "", "source_file": file_name, "row_number": row_num + 2, "notes": "missing job or item_id"})
            continue

        item_id   = str(item_id)
        date      = get_value(row_vals, DATE_COL)
        cust_val  = get_value(row_vals, CUST_COL)
        requester = get_value(row_vals, REQ_COL)
        dept      = get_value(row_vals, DEPT_COL)
        quantity  = safe_decimal(get_value(row_vals, QTY_COL))

        if quantity is None:
            print(f"[WARN] No quantity — stock {stock_id} item {item_id} (row saved with null qty)")

        item = caches["items"].get(item_id)
        if item is None:
            print(f"[CRIT] No matching item {item_id} for outstock {stock_id}")
            writer.writerow({"stock_id": stock_id or "", "item": item_id, "source_file": file_name, "row_number": row_num + 2, "notes": "no matching item in DB"})
            continue

        job_str = str(job_val)

        customer = None
        if cust_val:
            customer, _ = get_or_create_cached(caches["customers"], Customer, "name", str(cust_val))

        # Accumulate item changes — deduped by code (only when quantity present)
        if quantity is not None:
            tracked = items_by_code.get(item_id, item)
            tracked.quantity = max(Decimal(0), (tracked.quantity or Decimal(0)) - quantity)
            tracked.outstock_number += 1
            items_by_code[item_id] = tracked

        pending_outstock.append(Outstock(
            stock_id=str(stock_id) if stock_id else "", item=item,
            stock_date=date, requester=requester or "",
            quantity=quantity or Decimal(0), department=dept,
            store_type=store_type, customer=customer,
        ))
        pending_jobs.append(job_str)

        if len(pending_outstock) >= FLUSH_EVERY:
            created = flush_stock_records(pending_outstock, Outstock)
            _assign_jobs(created, pending_jobs)
            pending_jobs.clear()
            flush_items(items_by_code, ITEM_FIELDS)
            progress[key] = row_num + 1
            save_progress(progress)
            print(f"[PROGRESS] {sheet_name} — row {row_num + 1}")

    # Final flush
    created = flush_stock_records(pending_outstock, Outstock)
    _assign_jobs(created, pending_jobs)
    pending_jobs.clear()
    flush_items(items_by_code, ITEM_FIELDS)
    progress.pop(key, None)
    save_progress(progress)
    print(f"[INFO] Outstock sheet complete: {sheet_name}")


# ---------------------------------------------------------------------------
# Job M2M helper
# ---------------------------------------------------------------------------

_job_cache = {}

def _get_or_create_job(job_id_str):
    """Get or create a Job by its string ID. Returns None if empty/invalid."""
    if job_id_str is None:
        return None
    job_id_str = str(job_id_str).strip()
    if not job_id_str or job_id_str == '.':
        return None
    # Strip trailing .0 from numeric strings (Excel float artifacts)
    if '.' in job_id_str:
        try:
            job_id_str = str(int(float(job_id_str)))
        except (ValueError, TypeError):
            pass
    if not job_id_str:
        return None
    if job_id_str in _job_cache:
        return _job_cache[job_id_str]
    job, _ = Job.objects.get_or_create(job_id=job_id_str)
    _job_cache[job_id_str] = job
    return job


def _assign_jobs(created_records, job_strings):
    """Assign M2M job relations for a batch of created stock records."""
    for record, job_str in zip(created_records, job_strings):
        if job_str is None:
            continue
        job = _get_or_create_job(job_str)
        if job is not None:
            record.job.add(job)


# ---------------------------------------------------------------------------
# Price helpers
# ---------------------------------------------------------------------------

def _min_price(new_price, item):
    if item.min_price is None or item.min_price < 0:
        return Decimal(str(new_price))
    return min(item.min_price, Decimal(str(new_price)))


def _max_price(new_price, item):
    if item.max_price is None or item.max_price < 0:
        return Decimal(str(new_price))
    return max(item.max_price, Decimal(str(new_price)))